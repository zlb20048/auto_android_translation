# -*- coding: utf-8 -*-

import sys
import io
import os
import subprocess
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from pathlib import Path
import difflib
import xml.dom.minidom

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 修改常量定义
REPO_URL = "ssh://zixiangliu@10.10.96.213:29418/projects/Pateo-APP-MainInteraction-Hyundai/manifest"
USERNAME = "zixiangliu"
BRANCH = "master"
TRANSLATION_EXCEL_PATH = "/home/zixiangliu/文档/文言/translation.xlsx"
BASE_DIR = Path("/work/11") / Path(REPO_URL.split('/')[-2])
TEMP_EXCEL_PATH = BASE_DIR / "temp_translations.xlsx"

# 在文件顶部添加这个新函数
def get_git_root(path):
    try:
        git_root = subprocess.check_output(['git', '-C', str(path), 'rev-parse', '--show-toplevel'], stderr=subprocess.DEVNULL)
        logging.info(f"git_root ---> {git_root}")
        return Path(git_root.decode('utf-8').strip())
    except subprocess.CalledProcessError:
        # 如果当前目录不是 git 仓库，尝试父目录
        parent = path.parent
        if parent == path:  # 已经到达文件系统根目录
            logging.warning(f"无法获取 git 根目录: {path}")
            return None
        return get_git_root(parent)

def clone_repo():
    logging.info("开始克隆代码仓库")
    try:
        BASE_DIR.mkdir(parents=True, exist_ok=True)
        os.chdir(BASE_DIR)
        env = os.environ.copy()
        env['LANG'] = 'en_US.UTF-8'
        env['LC_ALL'] = 'en_US.UTF-8'

        # 使用 Popen 执行命令
        process = subprocess.Popen(
            ["repo", "init", "-u", REPO_URL, "-b", BRANCH],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env,
            universal_newlines=True
        )
        stdout, stderr = process.communicate()

        if process.returncode != 0:
            raise subprocess.CalledProcessError(process.returncode, process.args, stdout, stderr)

        logging.info(stdout)
        if stderr:
            logging.warning(stderr)

        # 执行 repo sync
        process = subprocess.Popen(
            ["repo", "sync"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env,
            universal_newlines=True
        )
        stdout, stderr = process.communicate()

        if process.returncode != 0:
            raise subprocess.CalledProcessError(process.returncode, process.args, stdout, stderr)

        logging.info(stdout)
        if stderr:
            logging.warning(stderr)

        logging.info("代码仓库克隆完成")
    except subprocess.CalledProcessError as e:
        logging.error(f"克隆仓库失败: {e}")
        logging.error(f"命令输出: {e.output}")
        logging.error(f"错误输出: {e.stderr}")
        raise
    except Exception as e:
        logging.error(f"克隆仓库时发生未知错误: {e}")
        raise

def process_strings_xml(xml_path):
    logging.info(f"处理 strings.xml 文件: {xml_path}")
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        strings = {}
        for string in root.findall('string'):
            name = string.get('name')
            value = string.text
            strings[name] = value
        return strings
    except ET.ParseError as e:
        logging.error(f"解析 XML 文件失败: {e}")
        return {}

def update_temp_excel(strings_dict, sheet_name, xml_path):
    logging.info(f"更新临时 Excel 表格，sheet: {sheet_name}")
    try:
        if not TEMP_EXCEL_PATH.exists():
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认创建的 sheet
        else:
            wb = load_workbook(TEMP_EXCEL_PATH)
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # 修改表头
        headers = ["Key", "Default", "FilePath"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # 更新数据
        for row, (key, value) in enumerate(strings_dict.items(), start=2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=str(xml_path))
        
        wb.save(TEMP_EXCEL_PATH)
        logging.info(f"临时 Excel 表格更新完成: {sheet_name}")
    except Exception as e:
        logging.error(f"更新临时 Excel 表格失败: {e}")

def match_translations(temp_df, main_excel):
    logging.info("开始匹配翻译")
    try:
        # 获取所有语言列
        language_columns = set()
        for sheet_name, main_df in main_excel.items():
            language_columns.update([col for col in main_df.columns if col.startswith('values-')])

        # 为每种语言创建新列
        for lang in language_columns:
            if lang not in temp_df.columns:
                temp_df[lang] = ''

        # 遍历主翻译文件中的所有 sheet
        for sheet_name, main_df in main_excel.items():
            logging.info(f"处理 sheet: {sheet_name}")
            
            # 确保 'values' 列存在
            if 'values' not in main_df.columns:
                logging.warning(f"Sheet '{sheet_name}' 中未找到 'values' 列，跳过")
                continue

            # 创建一个字典，用于快速查找
            translation_map = {row['values']: row.to_dict() for _, row in main_df.iterrows()}

            # 使用映射更新 temp_df
            for idx, row in temp_df.iterrows():
                if row['Default'] in translation_map:
                    matched_row = translation_map[row['Default']]
                    for lang in language_columns:
                        if lang in matched_row and pd.notna(matched_row[lang]):
                            if pd.isna(temp_df.at[idx, lang]) or temp_df.at[idx, lang] == '':
                                temp_df.at[idx, lang] = matched_row[lang]

        logging.info("翻译匹配完成")
        return temp_df
    except Exception as e:
        logging.error(f"匹配翻译失败: {e}")
        return temp_df

def update_strings_xml(updated_excel):
    logging.info("开始更新 strings.xml 文件")
    try:
        for root, dirs, files in os.walk(BASE_DIR):
            if 'strings.xml' in files:
                xml_path = Path(root) / 'strings.xml'
                lang_code = xml_path.parent.name
                
                tree = ET.parse(xml_path)
                root = tree.getroot()
                
                for string in root.findall('string'):
                    name = string.get('name')
                    if name in updated_excel.index:
                        new_value = updated_excel.loc[name, lang_code]
                        if pd.notna(new_value):
                            string.text = str(new_value)
                
                tree.write(xml_path, encoding='utf-8', xml_declaration=True)
                logging.info(f"更新了 strings.xml 文件: {xml_path}")
    except Exception as e:
        logging.error(f"更新 strings.xml 文件失败: {e}")

def find_most_similar_sheet(sheet_name, available_sheets):
    """
    找到最相似的 sheet 名称
    """
    if sheet_name in available_sheets:
        return sheet_name
    
    similarities = [(s, difflib.SequenceMatcher(None, sheet_name, s).ratio()) for s in available_sheets]
    most_similar = max(similarities, key=lambda x: x[1])
    
    if most_similar[1] > 0.6:  # 设置一个相似度阈值
        logging.info(f"找到相似的 sheet 名称: '{sheet_name}' -> '{most_similar[0]}'")
        return most_similar[0]
    else:
        logging.warning(f"未找到与 '{sheet_name}' 相似的 sheet")
        return None

def update_project_strings_xml(temp_excel, base_dir):
    logging.info("开始更新项目中的 strings.xml 文件")
    try:
        for sheet_name, df in temp_excel.items():
            # 只处理非默认语言列
            lang_columns = [col for col in df.columns if col.startswith('values-')]

            for lang_column in lang_columns:
                # 按文件路径分组
                grouped = df.groupby('FilePath')
                
                for file_path, group in grouped:
                    xml_path = Path(file_path).parent.parent / lang_column / 'strings.xml'

                    # 确保目录存在
                    xml_path.parent.mkdir(parents=True, exist_ok=True)

                    # 创建或更新 XML 文件
                    root = ET.Element("resources")

                    # 添加或更新字符串
                    updated = False
                    for _, row in group.iterrows():
                        if pd.notna(row[lang_column]):
                            string_elem = ET.SubElement(root, "string", name=row['Key'])
                            string_elem.text = str(row[lang_column])
                            updated = True

                    if updated:
                        # 使用 xml.dom.minidom 来格式化 XML
                        rough_string = ET.tostring(root, 'utf-8')
                        reparsed = xml.dom.minidom.parseString(rough_string)
                        pretty_xml = reparsed.toprettyxml(indent="    ")
                        
                        # 移除空行
                        pretty_xml = '\n'.join([line for line in pretty_xml.split('\n') if line.strip()])
                        
                        # 写入格式化后的 XML
                        with open(xml_path, 'w', encoding='utf-8') as f:
                            f.write(pretty_xml)
                        logging.info(f"更新或创建了 strings.xml 文件: {xml_path}")
                    else:
                        logging.info(f"没有需要更新的内容，跳过文件: {xml_path}")

        logging.info("所有 strings.xml 文件更新完成")
    except Exception as e:
        logging.error(f"更新项目 strings.xml 文件失败: {e}")
        raise

def main():
    try:
        # 1-6. 克隆代码仓库
        clone_repo()
        
        # 7-8. 处理 strings.xml 文件并更新到临时 Excel
        for xml_path in BASE_DIR.rglob('values/strings.xml'):
            strings_dict = process_strings_xml(xml_path)
            
            # 获取 git 根目录
            git_root = get_git_root(xml_path.parent)
            if git_root is None:
                logging.warning(f"跳过文件 {xml_path}，无法确定 git 根目录")
                continue
            
            # 使用相对于 BASE_DIR 的路径作为 sheet 名称
            relative_path = xml_path.relative_to(BASE_DIR)
            logging.info(f"relative_path --> {relative_path}")
            
            # 获取项目文件夹名（应该是 PateoSystemUI）
            project_folder = relative_path.parts[2]  # 假设路径结构是 Pateo-APP-MainInteraction-Hyundai/packages/app/PateoSystemUI/...
            internal_folder = relative_path.parts[3]
            # 构建新的 sheet_name
            sheet_name = f"{project_folder}_{internal_folder}"
            
            update_temp_excel(strings_dict, sheet_name, xml_path)
            logging.info(f"处理了文件: {xml_path}, sheet 名称: {sheet_name}")
        
        logging.info(f"临时 Excel 文件已创建: {TEMP_EXCEL_PATH}")
        
        # 9. 从 TRANSLATION_EXCEL_PATH 匹配翻译
        try:
            main_excel = pd.read_excel(TRANSLATION_EXCEL_PATH, sheet_name=None, engine='openpyxl')
        except Exception as e:
            logging.error(f"读取主翻译 Excel 文件失败: {e}")
            raise

        try:
            temp_excel = pd.read_excel(TEMP_EXCEL_PATH, sheet_name=None, engine='openpyxl')
        except Exception as e:
            logging.error(f"读取临时 Excel 文件失败: {e}")
            raise
        
        for sheet_name, df in temp_excel.items():
            updated_df = match_translations(df, main_excel)
            temp_excel[sheet_name] = updated_df
        
        # 保存更新后的临时 Excel 文件
        try:
            with pd.ExcelWriter(TEMP_EXCEL_PATH, engine='openpyxl') as writer:
                for sheet_name, df in temp_excel.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            logging.info("翻译匹配完成，临时 Excel 文件已更新")
        except Exception as e:
            logging.error(f"保存更新后的时 Excel 文件失败: {e}")
            raise

        # 10. 更新项目中的 strings.xml 文件
        update_project_strings_xml(temp_excel, BASE_DIR)

        logging.info("所有操作已完成")
    except Exception as e:
        logging.error(f"程序执行过程中发生错误: {e}")
        raise

if __name__ == "__main__":
    main()