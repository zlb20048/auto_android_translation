import pandas as pd
from openpyxl import Workbook, load_workbook
import logging

def update_temp_excel(strings_dict, sheet_name, xml_path, temp_excel_path):
    logging.info(f"更新临时 Excel 表格，sheet: {sheet_name}")
    try:
        if not temp_excel_path.exists():
            wb = Workbook()
            wb.remove(wb.active)
        else:
            wb = load_workbook(temp_excel_path)
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        headers = ["Key", "Default", "FilePath"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        for row, (key, value) in enumerate(strings_dict.items(), start=2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=str(xml_path))
        
        wb.save(temp_excel_path)
        logging.info(f"临时 Excel 表格更新完成: {sheet_name}")
    except Exception as e:
        logging.error(f"更新临时 Excel 表格失败: {e}")

def match_translations(temp_df, main_excel):
    logging.info("开始匹配翻译")
    try:
        language_columns = set()
        for sheet_name, main_df in main_excel.items():
            language_columns.update([col for col in main_df.columns if col.startswith('values-')])

        for lang in language_columns:
            if lang not in temp_df.columns:
                temp_df[lang] = ''

        for sheet_name, main_df in main_excel.items():
            logging.info(f"处理 sheet: {sheet_name}")
            
            if 'values' not in main_df.columns:
                logging.warning(f"Sheet '{sheet_name}' 中未找到 'values' 列，跳过")
                continue

            translation_map = {row['values']: row.to_dict() for _, row in main_df.iterrows()}

            for idx, row in temp_df.iterrows():
                if row['Default'] in translation_map:
                    matched_row = translation_map[row['Default']]
                    for lang in language_columns:
                        if lang in matched_row and pd.notna(matched_row[lang]):
                            if pd.isna(temp_df.at[idx, lang]) or temp_df.at[idx, lang] == '':
                                temp_df.at[idx, lang] = matched_row[lang]

        logging.info("翻译匹配完成")
        return temp_df
    except Exception as e:
        logging.error(f"匹配翻译失败: {e}")
        return temp_df
